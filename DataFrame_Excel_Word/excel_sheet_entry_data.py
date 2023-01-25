from tkinter import *
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib

cor = '#99B2F5'
cor2 = '#FFEF3A'
cor3 = 'black'
cor4 = 'white'

janela = Tk()

janela.title('Excel Dados')
janela.geometry('800x550')
janela.resizable(False, False)
janela.configure(bg=f'{cor}')

op = int(input('Informe: '
               '\n1 para preencher dados em uma linha completa'
               '\n2 para sobrescrever uma linha'
               '\n3 para escolher qual linha cada palavra irá'
               '\nDigite sua opção: '))

file = pathlib.Path('Palavras em Inglês.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "A"
    sheet['B1'] = "B"
    sheet['C1'] = "C"
    sheet['D1'] = "D"
    sheet['E1'] = "E"
    sheet['F1'] = "F"
    sheet['G1'] = "G"
    sheet['H1'] = "H"
    sheet['I1'] = "I"
    sheet['J1'] = "J"
    sheet['K1'] = "K"
    sheet['L1'] = "L"
    sheet['M1'] = "M"
    sheet['N1'] = "N"
    sheet['O1'] = "O"
    sheet['P1'] = "P"
    sheet['K1'] = "Q"
    sheet['R1'] = "R"
    sheet['S1'] = "S"
    sheet['T1'] = "T"
    sheet['U1'] = "U"
    sheet['V1'] = "V"
    sheet['W1'] = "W"
    sheet['X1'] = "X"
    sheet['Y1'] = "Y"
    sheet['Z1'] = "Z"

    file.save('Palavras em inglês.xlsx')


def enviar():
    a = aValue.get()
    b = bValue.get()
    c = cValue.get()
    d = dValue.get()
    e = eValue.get()
    f = fValue.get()
    g = gValue.get()
    h = hValue.get()
    i = iValue.get()
    j = jValue.get()
    k = kValue.get()
    l = lValue.get()
    m = mValue.get()
    n = nValue.get()
    o = oValue.get()
    p = pValue.get()
    q = qValue.get()
    r = rValue.get()
    s = sValue.get()
    t = tValue.get()
    u = uValue.get()
    v = vValue.get()
    w = wValue.get()
    x = xValue.get()
    y = yValue.get()
    z = zValue.get()

    file = openpyxl.load_workbook('Palavras em Inglês.xlsx')
    sheet = file.active

    if op == 1:
        sheet.cell(column=1, row=sheet.max_row + 1, value=a)
        sheet.cell(column=2, row=sheet.max_row, value=b)
        sheet.cell(column=3, row=sheet.max_row, value=c)
        sheet.cell(column=4, row=sheet.max_row, value=d)
        sheet.cell(column=5, row=sheet.max_row, value=e)
        sheet.cell(column=6, row=sheet.max_row, value=f)
        sheet.cell(column=7, row=sheet.max_row, value=g)
        sheet.cell(column=8, row=sheet.max_row, value=h)
        sheet.cell(column=9, row=sheet.max_row, value=i)
        sheet.cell(column=10, row=sheet.max_row, value=j)
        sheet.cell(column=11, row=sheet.max_row, value=k)
        sheet.cell(column=12, row=sheet.max_row, value=l)
        sheet.cell(column=13, row=sheet.max_row, value=m)
        sheet.cell(column=14, row=sheet.max_row, value=n)
        sheet.cell(column=15, row=sheet.max_row, value=o)
        sheet.cell(column=16, row=sheet.max_row, value=p)
        sheet.cell(column=17, row=sheet.max_row, value=q)
        sheet.cell(column=18, row=sheet.max_row, value=r)
        sheet.cell(column=19, row=sheet.max_row, value=s)
        sheet.cell(column=20, row=sheet.max_row, value=t)
        sheet.cell(column=21, row=sheet.max_row, value=u)
        sheet.cell(column=22, row=sheet.max_row, value=v)
        sheet.cell(column=23, row=sheet.max_row, value=w)
        sheet.cell(column=24, row=sheet.max_row, value=x)
        sheet.cell(column=25, row=sheet.max_row, value=y)
        sheet.cell(column=26, row=sheet.max_row, value=z)
    elif op == 2:
        sheet.cell(column=1, row=sheet.max_row, value=a)
        sheet.cell(column=2, row=sheet.max_row, value=b)
        sheet.cell(column=3, row=sheet.max_row, value=c)
        sheet.cell(column=4, row=sheet.max_row, value=d)
        sheet.cell(column=5, row=sheet.max_row, value=e)
        sheet.cell(column=6, row=sheet.max_row, value=f)
        sheet.cell(column=7, row=sheet.max_row, value=g)
        sheet.cell(column=8, row=sheet.max_row, value=h)
        sheet.cell(column=9, row=sheet.max_row, value=i)
        sheet.cell(column=10, row=sheet.max_row, value=j)
        sheet.cell(column=11, row=sheet.max_row, value=k)
        sheet.cell(column=12, row=sheet.max_row, value=l)
        sheet.cell(column=13, row=sheet.max_row, value=m)
        sheet.cell(column=14, row=sheet.max_row, value=n)
        sheet.cell(column=15, row=sheet.max_row, value=o)
        sheet.cell(column=16, row=sheet.max_row, value=p)
        sheet.cell(column=17, row=sheet.max_row, value=q)
        sheet.cell(column=18, row=sheet.max_row, value=r)
        sheet.cell(column=19, row=sheet.max_row, value=s)
        sheet.cell(column=20, row=sheet.max_row, value=t)
        sheet.cell(column=21, row=sheet.max_row, value=u)
        sheet.cell(column=22, row=sheet.max_row, value=v)
        sheet.cell(column=23, row=sheet.max_row, value=w)
        sheet.cell(column=24, row=sheet.max_row, value=x)
        sheet.cell(column=25, row=sheet.max_row, value=y)
        sheet.cell(column=26, row=sheet.max_row, value=z)
    elif op == 3:
        aLinha = int(input('Informe em qual linha será preenchida a palavra que começa com A: '))
        bLinha = int(input('Informe em qual linha será preenchida a palavra que começa com B: '))
        cLinha = int(input('Informe em qual linha será preenchida a palavra que começa com C: '))
        dLinha = int(input('Informe em qual linha será preenchida a palavra que começa com D: '))
        eLinha = int(input('Informe em qual linha será preenchida a palavra que começa com E: '))
        fLinha = int(input('Informe em qual linha será preenchida a palavra que começa com F: '))
        gLinha = int(input('Informe em qual linha será preenchida a palavra que começa com G: '))
        hLinha = int(input('Informe em qual linha será preenchida a palavra que começa com H: '))
        iLinha = int(input('Informe em qual linha será preenchida a palavra que começa com I: '))
        jLinha = int(input('Informe em qual linha será preenchida a palavra que começa com J: '))
        kLinha = int(input('Informe em qual linha será preenchida a palavra que começa com K: '))
        lLinha = int(input('Informe em qual linha será preenchida a palavra que começa com L: '))
        mLinha = int(input('Informe em qual linha será preenchida a palavra que começa com M: '))
        nLinha = int(input('Informe em qual linha será preenchida a palavra que começa com N: '))
        oLinha = int(input('Informe em qual linha será preenchida a palavra que começa com O: '))
        pLinha = int(input('Informe em qual linha será preenchida a palavra que começa com P: '))
        qLinha = int(input('Informe em qual linha será preenchida a palavra que começa com Q: '))
        rLinha = int(input('Informe em qual linha será preenchida a palavra que começa com R: '))
        sLinha = int(input('Informe em qual linha será preenchida a palavra que começa com S: '))
        tLinha = int(input('Informe em qual linha será preenchida a palavra que começa com T: '))
        uLinha = int(input('Informe em qual linha será preenchida a palavra que começa com U: '))
        vLinha = int(input('Informe em qual linha será preenchida a palavra que começa com V: '))
        wLinha = int(input('Informe em qual linha será preenchida a palavra que começa com W: '))
        xLinha = int(input('Informe em qual linha será preenchida a palavra que começa com X '))
        yLinha = int(input('Informe em qual linha será preenchida a palavra que começa com Y: '))
        zLinha = int(input('Informe em qual linha será preenchida a palavra que começa com Z: '))

        sheet.cell(column=1, row=aLinha, value=a)
        sheet.cell(column=2, row=bLinha, value=b)
        sheet.cell(column=3, row=cLinha, value=c)
        sheet.cell(column=4, row=dLinha, value=d)
        sheet.cell(column=5, row=eLinha, value=e)
        sheet.cell(column=6, row=fLinha, value=f)
        sheet.cell(column=7, row=gLinha, value=g)
        sheet.cell(column=8, row=hLinha, value=h)
        sheet.cell(column=9, row=iLinha, value=i)
        sheet.cell(column=10, row=jLinha, value=j)
        sheet.cell(column=11, row=kLinha, value=k)
        sheet.cell(column=12, row=lLinha, value=l)
        sheet.cell(column=13, row=mLinha, value=m)
        sheet.cell(column=14, row=nLinha, value=n)
        sheet.cell(column=15, row=oLinha, value=o)
        sheet.cell(column=16, row=pLinha, value=p)
        sheet.cell(column=17, row=qLinha, value=q)
        sheet.cell(column=18, row=rLinha, value=r)
        sheet.cell(column=19, row=sLinha, value=s)
        sheet.cell(column=20, row=tLinha, value=t)
        sheet.cell(column=21, row=uLinha, value=u)
        sheet.cell(column=22, row=vLinha, value=v)
        sheet.cell(column=23, row=wLinha, value=w)
        sheet.cell(column=24, row=xLinha, value=x)
        sheet.cell(column=25, row=yLinha, value=y)
        sheet.cell(column=26, row=zLinha, value=z)

    else:
        print('\n\033[31mOpção inválida!\033[m')

    file.save(r'Palavras em Inglês.xlsx')

    messagebox.showinfo('cool', 'datas added!')

    aValue.set('')
    bValue.set('')
    cValue.set('')
    dValue.set('')
    eValue.set('')
    fValue.set('')
    gValue.set('')
    hValue.set('')
    iValue.set('')
    jValue.set('')
    kValue.set('')
    lValue.set('')
    mValue.set('')
    nValue.set('')
    oValue.set('')
    pValue.set('')
    qValue.set('')
    rValue.set('')
    sValue.set('')
    tValue.set('')
    uValue.set('')
    vValue.set('')
    wValue.set('')
    xValue.set('')
    yValue.set('')
    zValue.set('')


def limpar():
    aValue.set('')
    bValue.set('')
    cValue.set('')
    dValue.set('')
    eValue.set('')
    fValue.set('')
    gValue.set('')
    hValue.set('')
    iValue.set('')
    jValue.set('')
    kValue.set('')
    lValue.set('')
    mValue.set('')
    nValue.set('')
    oValue.set('')
    pValue.set('')
    qValue.set('')
    rValue.set('')
    sValue.set('')
    tValue.set('')
    uValue.set('')
    vValue.set('')
    wValue.set('')
    xValue.set('')
    yValue.set('')
    zValue.set('')


# Icon:
icon = PhotoImage(file='excel_logo.png')
janela.iconphoto(False, icon)

# Heading:
Label(janela, text='Preencha os campos abaixo:', font="arial 13", bg=f'{cor}', fg=f'{cor2}').place(x=30, y=30)

# Label:
Label(janela, text='A', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=30, y=100)
Label(janela, text='B', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=30, y=150)
Label(janela, text='C', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=30, y=200)
Label(janela, text='D', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=30, y=250)
Label(janela, text='E', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=30, y=300)
Label(janela, text='F', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=30, y=350)
Label(janela, text='G', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=30, y=400)
Label(janela, text='H', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=30, y=450)
Label(janela, text='I', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=30, y=500)
Label(janela, text='J', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=230, y=100)
Label(janela, text='K', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=230, y=150)
Label(janela, text='L', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=230, y=200)
Label(janela, text='M', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=230, y=250)
Label(janela, text='N', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=230, y=300)
Label(janela, text='O', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=230, y=350)
Label(janela, text='P', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=230, y=400)
Label(janela, text='Q', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=230, y=450)
Label(janela, text='R', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=230, y=500)
Label(janela, text='S', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=430, y=100)
Label(janela, text='T', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=430, y=150)
Label(janela, text='U', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=430, y=200)
Label(janela, text='V', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=430, y=250)
Label(janela, text='W', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=430, y=300)
Label(janela, text='X', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=430, y=350)
Label(janela, text='Y', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=430, y=400)
Label(janela, text='Z', font=20, bg=f'{cor}', fg=f'{cor2}').place(x=430, y=450)

# Entry:

aValue = StringVar()
bValue = StringVar()
cValue = StringVar()
dValue = StringVar()
eValue = StringVar()
fValue = StringVar()
gValue = StringVar()
hValue = StringVar()
iValue = StringVar()
jValue = StringVar()
kValue = StringVar()
lValue = StringVar()
mValue = StringVar()
nValue = StringVar()
oValue = StringVar()
pValue = StringVar()
qValue = StringVar()
rValue = StringVar()
sValue = StringVar()
tValue = StringVar()
uValue = StringVar()
vValue = StringVar()
wValue = StringVar()
xValue = StringVar()
yValue = StringVar()
zValue = StringVar()

aEntry = Entry(janela, textvariable=aValue, width=18, bd=2, font=16)
bEntry = Entry(janela, textvariable=bValue, width=18, bd=2, font=16)
cEntry = Entry(janela, textvariable=cValue, width=18, bd=2, font=16)
dEntry = Entry(janela, textvariable=dValue, width=18, bd=2, font=16)
eEntry = Entry(janela, textvariable=eValue, width=18, bd=2, font=16)
fEntry = Entry(janela, textvariable=fValue, width=18, bd=2, font=16)
gEntry = Entry(janela, textvariable=gValue, width=18, bd=2, font=16)
hEntry = Entry(janela, textvariable=hValue, width=18, bd=2, font=16)
iEntry = Entry(janela, textvariable=iValue, width=18, bd=2, font=16)
jEntry = Entry(janela, textvariable=jValue, width=18, bd=2, font=16)
kEntry = Entry(janela, textvariable=kValue, width=18, bd=2, font=16)
lEntry = Entry(janela, textvariable=lValue, width=18, bd=2, font=16)
mEntry = Entry(janela, textvariable=mValue, width=18, bd=2, font=16)
nEntry = Entry(janela, textvariable=nValue, width=18, bd=2, font=16)
oEntry = Entry(janela, textvariable=oValue, width=18, bd=2, font=16)
pEntry = Entry(janela, textvariable=pValue, width=18, bd=2, font=16)
qEntry = Entry(janela, textvariable=qValue, width=18, bd=2, font=16)
rEntry = Entry(janela, textvariable=rValue, width=18, bd=2, font=16)
sEntry = Entry(janela, textvariable=sValue, width=18, bd=2, font=16)
tEntry = Entry(janela, textvariable=tValue, width=18, bd=2, font=16)
uEntry = Entry(janela, textvariable=uValue, width=18, bd=2, font=16)
vEntry = Entry(janela, textvariable=vValue, width=18, bd=2, font=16)
wEntry = Entry(janela, textvariable=wValue, width=18, bd=2, font=16)
xEntry = Entry(janela, textvariable=xValue, width=18, bd=2, font=16)
yEntry = Entry(janela, textvariable=yValue, width=18, bd=2, font=16)
zEntry = Entry(janela, textvariable=zValue, width=18, bd=2, font=16)

aEntry.place(x=50, y=100)
bEntry.place(x=50, y=150)
cEntry.place(x=50, y=200)
dEntry.place(x=50, y=250)
eEntry.place(x=50, y=300)
fEntry.place(x=50, y=350)
gEntry.place(x=50, y=400)
hEntry.place(x=50, y=450)
iEntry.place(x=50, y=500)
jEntry.place(x=250, y=100)
kEntry.place(x=250, y=150)
lEntry.place(x=250, y=200)
mEntry.place(x=250, y=250)
nEntry.place(x=250, y=300)
oEntry.place(x=250, y=350)
pEntry.place(x=250, y=400)
qEntry.place(x=250, y=450)
rEntry.place(x=250, y=500)
sEntry.place(x=450, y=100)
tEntry.place(x=450, y=150)
uEntry.place(x=450, y=200)
vEntry.place(x=450, y=250)
wEntry.place(x=450, y=300)
xEntry.place(x=450, y=350)
yEntry.place(x=450, y=400)
zEntry.place(x=450, y=450)

# Button:

Button(janela, text='Enviar', bg=f'{cor3}', fg=f"{cor4}", width=15, height=2, command=enviar).place(x=650, y=150)
Button(janela, text='Limpar', bg=f'{cor3}', fg=f"{cor4}", width=15, height=2, command=limpar).place(x=650, y=250)
Button(janela, text='Sair', bg=f'{cor3}', fg=f"{cor4}", width=15, height=2,
       command=lambda: janela.destroy()).place(x=650, y=350)

janela.mainloop()
